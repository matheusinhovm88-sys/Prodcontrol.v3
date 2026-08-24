import os
import json
import re
import unicodedata
from flask import Flask, jsonify, request
from flask_cors import CORS
import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl

app = Flask(__name__)
CORS(app)

# A chave NUNCA fica no código. Ela vem de uma variável de ambiente configurada
# no Vercel (Settings -> Environment Variables), com o conteúdo COMPLETO do
# arquivo JSON baixado do Firebase (Google Cloud) — sem editar nada nele.
if not firebase_admin._apps:
    chave_json = os.environ.get("FIREBASE_SERVICE_ACCOUNT_JSON")
    if not chave_json:
        raise RuntimeError(
            "Variável de ambiente FIREBASE_SERVICE_ACCOUNT_JSON não encontrada. "
            "Configure no Vercel em Settings -> Environment Variables."
        )
    # Blindagem contra copiar/colar pelo celular: se sobrar algum espaço, quebra de
    # linha ou texto extra antes/depois do JSON, pegamos só o trecho entre a
    # primeira "{" e a última "}" antes de tentar interpretar.
    inicio = chave_json.find("{")
    fim = chave_json.rfind("}")
    if inicio == -1 or fim == -1 or fim < inicio:
        raise RuntimeError(
            "FIREBASE_SERVICE_ACCOUNT_JSON não parece conter um JSON válido "
            "(não encontrei '{' e '}'). Revise o valor colado no Vercel."
        )
    try:
        credenciais_dict = json.loads(chave_json[inicio:fim + 1])
    except json.JSONDecodeError as erro:
        raise RuntimeError(
            f"FIREBASE_SERVICE_ACCOUNT_JSON contém JSON inválido: {erro}. "
            "Apague o valor no Vercel e cole novamente, só uma vez, o conteúdo "
            "exato do arquivo .json baixado do Google Cloud."
        ) from erro
    cred = credentials.Certificate(credenciais_dict)
    default_app = firebase_admin.initialize_app(cred)
else:
    default_app = firebase_admin.get_app()

db = firestore.client(app=default_app)


@app.route('/')
def home():
    return jsonify({"status": "API rodando com sucesso e conectada ao Firebase!"})


# Nome do arquivo da planilha, sempre o mesmo, commitado na raiz do repositório.
# Toda vez que a chefe manda uma nova, você substitui esse arquivo no GitHub.
CAMINHO_PLANILHA = os.path.join(os.path.dirname(__file__), "ordens.xlsx")

# Cada variante de nome de coluna que já vimos aparecer na planilha, já sem
# acento/maiúscula/pontuação/espaço (normalizado). Se a chefe mudar o nome de
# um jeito novo, é só adicionar a variante nova na lista certa aqui embaixo.
VARIANTES_COLUNA = {
    "lote": ["lote"],
    "ordem": ["ordem"],
    "qtde": ["qtde", "qtd", "quantidade"],
    "coditem": ["coditem", "coditens", "coditemcod", "codigoitem", "iditem", "codigo"],
    "descricao": ["descricao", "descr", "desc"],
    "unidade": ["unidade", "und"],
    "mascara": ["mascara", "mask"],
}


def normalizar_texto(texto):
    """minúsculo, sem acento, sem espaço/pontuação — pra comparar nomes de coluna com folga."""
    texto = str(texto or "")
    texto = unicodedata.normalize("NFD", texto).encode("ascii", "ignore").decode("utf-8")
    return re.sub(r"[^a-z0-9]", "", texto.lower())


def mapear_colunas(linha_cabecalho):
    """Descobre qual coluna (índice) corresponde a cada campo que a gente precisa."""
    normalizado_por_indice = {i: normalizar_texto(v) for i, v in enumerate(linha_cabecalho)}
    mapa = {}
    for campo, variantes in VARIANTES_COLUNA.items():
        for indice, nome_normalizado in normalizado_por_indice.items():
            if nome_normalizado in variantes:
                mapa[campo] = indice
                break
    return mapa


def verificar_autorizacao_cron():
    """Confere o segredo do Cron Job do Vercel, se estiver configurado."""
    segredo_esperado = os.environ.get("CRON_SECRET")
    if not segredo_esperado:
        return True  # nenhum segredo configurado ainda — não bloqueia (configure depois por segurança)
    cabecalho = request.headers.get("Authorization", "")
    return cabecalho == f"Bearer {segredo_esperado}"


@app.route('/api/importar')
def importar_ordens():
    if not verificar_autorizacao_cron():
        return jsonify({"erro": "Não autorizado."}), 401

    empresa_id = os.environ.get("PRODCONTROL_EMPRESA_ID")
    if not empresa_id:
        return jsonify({"erro": "Variável PRODCONTROL_EMPRESA_ID não configurada no Vercel."}), 500

    if not os.path.exists(CAMINHO_PLANILHA):
        return jsonify({"erro": f"Arquivo '{os.path.basename(CAMINHO_PLANILHA)}' não encontrado no repositório."}), 404

    try:
        pasta_trabalho = openpyxl.load_workbook(CAMINHO_PLANILHA, data_only=True)
        planilha = pasta_trabalho[pasta_trabalho.sheetnames[0]]  # sempre a primeira aba, não importa o nome dela
        linhas = list(planilha.iter_rows(values_only=True))
    except Exception as erro:
        return jsonify({"erro": f"Não consegui ler a planilha: {erro}"}), 400

    if not linhas:
        return jsonify({"erro": "Planilha vazia."}), 400

    mapa = mapear_colunas(linhas[0])
    faltando = [c for c in ("ordem", "coditem") if c not in mapa]
    if faltando:
        return jsonify({
            "erro": f"Não encontrei a(s) coluna(s) obrigatória(s) {faltando} no cabeçalho da planilha.",
            "cabecalho_lido": list(linhas[0])
        }), 400

    importados = 0
    ignorados = 0
    erros_linha = []
    lote_gravacao = db.batch()
    operacoes_no_lote = 0

    for numero_linha, linha in enumerate(linhas[1:], start=2):
        try:
            ordem = linha[mapa["ordem"]] if mapa.get("ordem") is not None else None
            if ordem is None or str(ordem).strip() == "":
                ignorados += 1
                continue

            doc_ordem = {
                "lote": str(linha[mapa["lote"]]) if "lote" in mapa and linha[mapa["lote"]] is not None else "",
                "qtde": linha[mapa["qtde"]] if "qtde" in mapa and linha[mapa["qtde"]] is not None else None,
                "coditem": str(linha[mapa["coditem"]]) if linha[mapa["coditem"]] is not None else "",
                "descricao": str(linha[mapa["descricao"]]) if "descricao" in mapa and linha[mapa["descricao"]] is not None else "",
                "unidade": str(linha[mapa["unidade"]]) if "unidade" in mapa and linha[mapa["unidade"]] is not None else "",
                # Máscara pode legitimamente vir vazia — não é motivo pra pular a linha.
                "mascara": str(linha[mapa["mascara"]]) if "mascara" in mapa and linha[mapa["mascara"]] is not None else "",
                "atualizadoEm": firestore.SERVER_TIMESTAMP,
            }

            ref = db.collection("empresas", empresa_id, "ordens").document(str(ordem).strip())
            lote_gravacao.set(ref, doc_ordem)
            operacoes_no_lote += 1
            importados += 1

            # O Firestore aceita no máximo 500 operações por lote.
            if operacoes_no_lote >= 450:
                lote_gravacao.commit()
                lote_gravacao = db.batch()
                operacoes_no_lote = 0

        except Exception as erro_linha:
            ignorados += 1
            erros_linha.append(f"linha {numero_linha}: {erro_linha}")

    if operacoes_no_lote > 0:
        lote_gravacao.commit()

    return jsonify({
        "status": "Importação concluída.",
        "ordens_importadas": importados,
        "linhas_ignoradas": ignorados,
        "detalhe_erros": erros_linha[:20]  # não manda uma lista infinita se der muito erro
    })


if __name__ == '__main__':
    app.run(debug=True)